"""
Suivi Chantiers — Flask + PostgreSQL
Authentification par sessions, rôles : admin | gestionnaire | readonly
L'admin peut gérer les utilisateurs et la configuration par défaut des suivi.
"""
import os, uuid, io
from functools import wraps
from flask import (Flask, jsonify, request, render_template,
                   send_file, session, redirect, url_for)
import psycopg2, psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-CHANGE-IN-PROD')

DATABASE_URL = os.environ.get(
    'DATABASE_URL',
    'postgresql://suivi:suivi@localhost:5432/suivi_chantiers'
)

# ── Connexion PostgreSQL ───────────────────────────────────────────────────────

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False
    return conn

def qall(conn, sql, p=()):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
        c.execute(sql, p); return [dict(r) for r in c.fetchall()]

def qone(conn, sql, p=()):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as c:
        c.execute(sql, p); r = c.fetchone(); return dict(r) if r else None

def run(conn, sql, p=()):
    with conn.cursor() as c: c.execute(sql, p)

# ── Initialisation BDD ────────────────────────────────────────────────────────

DDL = """
CREATE EXTENSION IF NOT EXISTS "pgcrypto";

CREATE TABLE IF NOT EXISTS users (
    id           UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    username     VARCHAR(100) UNIQUE NOT NULL,
    email        VARCHAR(255),
    password_hash VARCHAR(255) NOT NULL,
    role         VARCHAR(20)  NOT NULL DEFAULT 'gestionnaire',
    is_active    BOOLEAN      NOT NULL DEFAULT TRUE,
    created_at   TIMESTAMP    NOT NULL DEFAULT NOW(),
    created_by   UUID REFERENCES users(id)
);

-- Catégories par défaut (éditables par l'admin)
CREATE TABLE IF NOT EXISTS default_categories (
    id        UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    nom       VARCHAR(255) NOT NULL,
    ordre     INTEGER NOT NULL DEFAULT 0,
    is_active BOOLEAN NOT NULL DEFAULT TRUE
);

-- Tâches par défaut par catégorie (éditables par l'admin)
CREATE TABLE IF NOT EXISTS default_tasks (
    id            UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    category_nom  VARCHAR(255) NOT NULL,
    nom           VARCHAR(255) NOT NULL,
    ordre         INTEGER NOT NULL DEFAULT 0,
    is_active     BOOLEAN NOT NULL DEFAULT TRUE
);

CREATE TABLE IF NOT EXISTS chantiers (
    id           UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    nom          VARCHAR(255) NOT NULL,
    adresse      TEXT,
    client       VARCHAR(255),
    logo_data    BYTEA,
    logo_mime    VARCHAR(100),
    date_debut   DATE,
    date_fin     DATE,
    commentaires TEXT,
    created_by   UUID REFERENCES users(id),
    created_at   TIMESTAMP NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS categories (
    id          UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    chantier_id UUID NOT NULL REFERENCES chantiers(id) ON DELETE CASCADE,
    nom         VARCHAR(255) NOT NULL,
    ordre       INTEGER NOT NULL DEFAULT 0,
    is_custom   BOOLEAN NOT NULL DEFAULT FALSE
);

CREATE TABLE IF NOT EXISTS taches (
    id           UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    categorie_id UUID NOT NULL REFERENCES categories(id) ON DELETE CASCADE,
    nom          VARCHAR(255) NOT NULL,
    etabli       SMALLINT NOT NULL DEFAULT 0,
    envoye       SMALLINT NOT NULL DEFAULT 0,
    valide       SMALLINT NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS planning (
    id          UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    chantier_id UUID NOT NULL REFERENCES chantiers(id) ON DELETE CASCADE,
    phase       VARCHAR(255) NOT NULL,
    date_debut  DATE NOT NULL,
    date_fin    DATE NOT NULL
);
"""

DEFAULT_CATEGORIES_SEED = [
    'Préparatoire', 'Montage', 'Maçonnerie', 'CE/Levée de réserve',
    'Démontage', 'Pose de SAS', 'Mise à disposition', 'Contrôle & Essai', 'Désamiantage'
]

DEFAULT_TASKS_SEED = [
    ('Préparatoire', 'Relevé de gaine'),
    ('Préparatoire', 'Demande Plan (Fournisseur)'),
    ('Préparatoire', 'Plan ascenseur (BE)'),
    ('Préparatoire', 'Commande Fournisseur'),
    ('Préparatoire', 'Demande Mise en FAB'),
    ('Préparatoire', 'Demande PGC'),
    ('Préparatoire', 'PPSPS Manei'),
    ('Préparatoire', 'Planning Prév fournisseur'),
    ('Préparatoire', 'Planning Prév BE/Client'),
    ('Préparatoire', 'Devis ST'),
    ('Préparatoire', 'Demande Agrément ST'),
    ('Préparatoire', 'Commande ST'),
    ('Préparatoire', 'PPSPS ST'),
    ('Préparatoire', 'Demande Date VIC'),
    ('Préparatoire', 'Commande Base de vie'),
    ('Préparatoire', 'Commande Container'),
]

def init_db():
    conn = get_db()
    try:
        with conn.cursor() as c:
            c.execute(DDL)
        # Admin par défaut si aucun utilisateur
        existing = qone(conn, "SELECT id FROM users LIMIT 1")
        if not existing:
            run(conn, """
                INSERT INTO users (username, email, password_hash, role)
                VALUES (%s, %s, %s, 'admin')
            """, ('admin', 'admin@local.fr', generate_password_hash('Admin1234!')))

        # Catégories par défaut si table vide
        if not qone(conn, "SELECT id FROM default_categories LIMIT 1"):
            for i, nom in enumerate(DEFAULT_CATEGORIES_SEED):
                run(conn, "INSERT INTO default_categories (nom, ordre) VALUES (%s, %s)", (nom, i))
        if not qone(conn, "SELECT id FROM default_tasks LIMIT 1"):
            for i, (cat, task) in enumerate(DEFAULT_TASKS_SEED):
                run(conn, "INSERT INTO default_tasks (category_nom, nom, ordre) VALUES (%s,%s,%s)",
                    (cat, task, i))
        conn.commit()
    finally:
        conn.close()

# ── Auth helpers ──────────────────────────────────────────────────────────────

def current_user():
    uid = session.get('user_id')
    if not uid: return None
    conn = get_db()
    try:
        return qone(conn, "SELECT id, username, role, is_active FROM users WHERE id=%s", (uid,))
    finally:
        conn.close()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        u = current_user()
        if not u or not u['is_active']:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Non authentifié'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        u = current_user()
        if not u or u['role'] != 'admin':
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Accès refusé'}), 403
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def readonly_check():
    """Retourne True si l'utilisateur courant est readonly (bloquer les mutations)."""
    u = current_user()
    return u and u['role'] == 'readonly'

# ── Pages HTML ────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET'])
def login_page():
    if current_user():
        return redirect('/')
    return render_template('login.html')

@app.route('/')
@login_required
def index():
    return render_template('index.html', user=current_user())

@app.route('/admin')
@admin_required
def admin_page():
    return render_template('admin.html', user=current_user())

# ── Auth API ──────────────────────────────────────────────────────────────────

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    d = request.json
    conn = get_db()
    try:
        u = qone(conn, "SELECT * FROM users WHERE username=%s AND is_active=TRUE",
                 (d.get('username',''),))
        if not u or not check_password_hash(u['password_hash'], d.get('password','')):
            return jsonify({'error': 'Identifiants incorrects'}), 401
        session['user_id'] = str(u['id'])
        return jsonify({'username': u['username'], 'role': u['role']})
    finally:
        conn.close()

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/auth/me')
@login_required
def api_me():
    u = current_user()
    return jsonify({'username': u['username'], 'role': u['role']})

# ── Admin : gestion utilisateurs ─────────────────────────────────────────────

@app.route('/api/admin/users')
@admin_required
def admin_get_users():
    conn = get_db()
    try:
        users = qall(conn, """
            SELECT u.id, u.username, u.email, u.role, u.is_active, u.created_at,
                   c.username AS created_by_name
            FROM users u
            LEFT JOIN users c ON c.id = u.created_by
            ORDER BY u.created_at
        """)
        for u in users:
            u['id']         = str(u['id'])
            u['created_at'] = u['created_at'].strftime('%d/%m/%Y') if u['created_at'] else ''
        return jsonify(users)
    finally:
        conn.close()

@app.route('/api/admin/users', methods=['POST'])
@admin_required
def admin_create_user():
    d = request.json
    if not d.get('username') or not d.get('password'):
        return jsonify({'error': 'username et password requis'}), 400
    conn = get_db()
    try:
        existing = qone(conn, "SELECT id FROM users WHERE username=%s", (d['username'],))
        if existing:
            return jsonify({'error': 'Ce nom d\'utilisateur existe déjà'}), 409
        uid = str(uuid.uuid4())
        run(conn, """
            INSERT INTO users (id, username, email, password_hash, role, created_by)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (uid, d['username'], d.get('email',''),
              generate_password_hash(d['password']),
              d.get('role','gestionnaire'),
              session.get('user_id')))
        conn.commit()
        return jsonify({'id': uid, 'username': d['username']}), 201
    finally:
        conn.close()

@app.route('/api/admin/users/<uid>', methods=['PUT'])
@admin_required
def admin_update_user(uid):
    d = request.json
    conn = get_db()
    try:
        if d.get('password'):
            run(conn, "UPDATE users SET password_hash=%s WHERE id=%s",
                (generate_password_hash(d['password']), uid))
        if 'role' in d:
            run(conn, "UPDATE users SET role=%s WHERE id=%s", (d['role'], uid))
        if 'email' in d:
            run(conn, "UPDATE users SET email=%s WHERE id=%s", (d['email'], uid))
        if 'is_active' in d:
            # Empêcher de désactiver son propre compte
            if str(uid) == str(session.get('user_id')) and not d['is_active']:
                return jsonify({'error': 'Impossible de désactiver votre propre compte'}), 400
            run(conn, "UPDATE users SET is_active=%s WHERE id=%s", (d['is_active'], uid))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/admin/users/<uid>', methods=['DELETE'])
@admin_required
def admin_delete_user(uid):
    if str(uid) == str(session.get('user_id')):
        return jsonify({'error': 'Impossible de supprimer votre propre compte'}), 400
    conn = get_db()
    try:
        run(conn, "DELETE FROM users WHERE id=%s", (uid,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

# ── Admin : configuration par défaut des suivis ───────────────────────────────

@app.route('/api/admin/defaults')
@admin_required
def admin_get_defaults():
    conn = get_db()
    try:
        cats  = qall(conn, "SELECT * FROM default_categories ORDER BY ordre")
        tasks = qall(conn, "SELECT * FROM default_tasks ORDER BY category_nom, ordre")
        for r in cats:  r['id'] = str(r['id'])
        for r in tasks: r['id'] = str(r['id'])
        return jsonify({'categories': cats, 'tasks': tasks})
    finally:
        conn.close()

@app.route('/api/admin/defaults/categories', methods=['POST'])
@admin_required
def admin_add_default_cat():
    d = request.json
    conn = get_db()
    try:
        max_ordre = qone(conn, "SELECT COALESCE(MAX(ordre),0)+1 AS n FROM default_categories")['n']
        cid = str(uuid.uuid4())
        run(conn, "INSERT INTO default_categories (id, nom, ordre) VALUES (%s,%s,%s)",
            (cid, d['nom'], max_ordre))
        conn.commit()
        return jsonify({'id': cid, 'nom': d['nom'], 'ordre': max_ordre, 'is_active': True}), 201
    finally:
        conn.close()

@app.route('/api/admin/defaults/categories/<cid>', methods=['PUT'])
@admin_required
def admin_update_default_cat(cid):
    d = request.json
    conn = get_db()
    try:
        if 'nom' in d:       run(conn, "UPDATE default_categories SET nom=%s WHERE id=%s",       (d['nom'], cid))
        if 'is_active' in d: run(conn, "UPDATE default_categories SET is_active=%s WHERE id=%s", (d['is_active'], cid))
        if 'ordre' in d:     run(conn, "UPDATE default_categories SET ordre=%s WHERE id=%s",     (d['ordre'], cid))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/admin/defaults/categories/<cid>', methods=['DELETE'])
@admin_required
def admin_delete_default_cat(cid):
    conn = get_db()
    try:
        cat = qone(conn, "SELECT nom FROM default_categories WHERE id=%s", (cid,))
        if cat:
            run(conn, "DELETE FROM default_tasks WHERE category_nom=%s", (cat['nom'],))
        run(conn, "DELETE FROM default_categories WHERE id=%s", (cid,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/admin/defaults/tasks', methods=['POST'])
@admin_required
def admin_add_default_task():
    d = request.json
    conn = get_db()
    try:
        max_o = qone(conn,
            "SELECT COALESCE(MAX(ordre),0)+1 AS n FROM default_tasks WHERE category_nom=%s",
            (d['category_nom'],))['n']
        tid = str(uuid.uuid4())
        run(conn, "INSERT INTO default_tasks (id, category_nom, nom, ordre) VALUES (%s,%s,%s,%s)",
            (tid, d['category_nom'], d['nom'], max_o))
        conn.commit()
        return jsonify({'id': tid, 'category_nom': d['category_nom'], 'nom': d['nom'],
                        'ordre': max_o, 'is_active': True}), 201
    finally:
        conn.close()

@app.route('/api/admin/defaults/tasks/<tid>', methods=['PUT'])
@admin_required
def admin_update_default_task(tid):
    d = request.json
    conn = get_db()
    try:
        if 'nom' in d:       run(conn, "UPDATE default_tasks SET nom=%s WHERE id=%s",       (d['nom'], tid))
        if 'is_active' in d: run(conn, "UPDATE default_tasks SET is_active=%s WHERE id=%s", (d['is_active'], tid))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/admin/defaults/tasks/<tid>', methods=['DELETE'])
@admin_required
def admin_delete_default_task(tid):
    conn = get_db()
    try:
        run(conn, "DELETE FROM default_tasks WHERE id=%s", (tid,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

# ── Chantiers ─────────────────────────────────────────────────────────────────

def compute_progress(conn, chantier_id):
    rows = qall(conn, """
        SELECT t.etabli, t.envoye, t.valide FROM taches t
        JOIN categories c ON t.categorie_id = c.id
        WHERE c.chantier_id = %s
    """, (chantier_id,))
    if not rows: return 0
    return round(sum(r['etabli']+r['envoye']+r['valide'] for r in rows) / (len(rows)*6) * 100)

def fmt_chantier(c, progress=0):
    return {
        'id':           str(c['id']),
        'nom':          c['nom'],
        'adresse':      c['adresse'],
        'client':       c['client'],
        'logoUrl':      f'/api/logos/{c["id"]}' if c.get('logo_data') else (c.get('logo_url') or ''),
        'dateDebut':    str(c['date_debut']) if c.get('date_debut') else '',
        'dateFin':      str(c['date_fin'])   if c.get('date_fin')   else '',
        'commentaires': c['commentaires'],
        'progress':     progress,
    }

@app.route('/api/chantiers')
@login_required
def get_chantiers():
    conn = get_db()
    try:
        chantiers = qall(conn, "SELECT * FROM chantiers ORDER BY nom")
        return jsonify([fmt_chantier(c, compute_progress(conn, c['id'])) for c in chantiers])
    finally:
        conn.close()

@app.route('/api/chantiers/<cid>')
@login_required
def get_chantier(cid):
    conn = get_db()
    try:
        c = qone(conn, "SELECT * FROM chantiers WHERE id=%s", (cid,))
        if not c: return jsonify({'error': 'Not found'}), 404
        cats = qall(conn, "SELECT * FROM categories WHERE chantier_id=%s ORDER BY ordre", (cid,))
        for cat in cats:
            tasks = qall(conn, "SELECT * FROM taches WHERE categorie_id=%s", (cat['id'],))
            score = sum(t['etabli']+t['envoye']+t['valide'] for t in tasks)
            maxs  = len(tasks)*6
            cat['id']         = str(cat['id'])
            cat['chantierId'] = str(cat['chantier_id'])
            cat['isCustom']   = cat['is_custom']
            cat['tasks']      = [fmt_tache(t) for t in tasks]
            cat['score']      = score
            cat['maxScore']   = maxs
            cat['totalCount'] = len(tasks)
            cat['progress']   = round(score/maxs*100) if maxs else 0
        result = fmt_chantier(c, compute_progress(conn, cid))
        result['categories'] = cats
        return jsonify(result)
    finally:
        conn.close()

@app.route('/api/chantiers', methods=['POST'])
@login_required
def create_chantier():
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    d = request.json
    cid = str(uuid.uuid4())
    conn = get_db()
    try:
        run(conn, """
            INSERT INTO chantiers (id, nom, adresse, client, date_debut, date_fin, commentaires, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (cid, d.get('nom',''), d.get('adresse',''), d.get('client',''),
              d.get('dateDebut') or None, d.get('dateFin') or None,
              d.get('commentaires',''), session.get('user_id')))

        # Catégories par défaut actives
        default_cats = qall(conn,
            "SELECT * FROM default_categories WHERE is_active=TRUE ORDER BY ordre")
        for i, dc in enumerate(default_cats):
            catid = str(uuid.uuid4())
            run(conn, """
                INSERT INTO categories (id, chantier_id, nom, ordre, is_custom)
                VALUES (%s,%s,%s,%s,FALSE)
            """, (catid, cid, dc['nom'], i))
            # Tâches par défaut pour cette catégorie
            tasks = qall(conn, """
                SELECT * FROM default_tasks
                WHERE category_nom=%s AND is_active=TRUE ORDER BY ordre
            """, (dc['nom'],))
            for t in tasks:
                run(conn, """
                    INSERT INTO taches (id, categorie_id, nom, etabli, envoye, valide)
                    VALUES (%s,%s,%s,0,0,0)
                """, (str(uuid.uuid4()), catid, t['nom']))

        conn.commit()
        return jsonify({'id': cid, **{k: d.get(k,'') for k in
                        ['nom','adresse','client','dateDebut','dateFin','commentaires']},
                        'logoUrl': '', 'progress': 0}), 201
    finally:
        conn.close()

@app.route('/api/chantiers/<cid>', methods=['PUT'])
@login_required
def update_chantier(cid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    d = request.json
    conn = get_db()
    try:
        run(conn, """
            UPDATE chantiers SET nom=%s, adresse=%s, client=%s,
            date_debut=%s, date_fin=%s, commentaires=%s WHERE id=%s
        """, (d.get('nom'), d.get('adresse'), d.get('client'),
              d.get('dateDebut') or None, d.get('dateFin') or None,
              d.get('commentaires'), cid))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/chantiers/<cid>', methods=['DELETE'])
@login_required
def delete_chantier(cid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    try:
        run(conn, "DELETE FROM chantiers WHERE id=%s", (cid,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

# ── Logo ──────────────────────────────────────────────────────────────────────

@app.route('/api/logos/<cid>')
@login_required
def get_logo(cid):
    conn = get_db()
    try:
        r = qone(conn, "SELECT logo_data, logo_mime FROM chantiers WHERE id=%s", (cid,))
    finally:
        conn.close()
    if not r or not r['logo_data']:
        return '', 404
    return send_file(io.BytesIO(bytes(r['logo_data'])), mimetype=r['logo_mime'] or 'image/png')

@app.route('/api/logos/<cid>', methods=['POST'])
@login_required
def upload_logo(cid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    if 'logo' not in request.files:
        return jsonify({'error': 'Aucun fichier'}), 400
    f = request.files['logo']
    data = f.read()
    if len(data) > 5*1024*1024:
        return jsonify({'error': 'Fichier trop volumineux (max 5 Mo)'}), 413
    conn = get_db()
    try:
        run(conn, "UPDATE chantiers SET logo_data=%s, logo_mime=%s WHERE id=%s",
            (psycopg2.Binary(data), f.mimetype or 'image/png', cid))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'url': f'/api/logos/{cid}'})

@app.route('/api/logos/<cid>', methods=['DELETE'])
@login_required
def delete_logo(cid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    try:
        run(conn, "UPDATE chantiers SET logo_data=NULL, logo_mime=NULL WHERE id=%s", (cid,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'success': True})

# ── Catégories ────────────────────────────────────────────────────────────────

@app.route('/api/categories', methods=['POST'])
@login_required
def create_categorie():
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    d = request.json
    conn = get_db()
    try:
        ordre = qone(conn, "SELECT COUNT(*) AS n FROM categories WHERE chantier_id=%s",
                     (d['chantierId'],))['n']
        catid = str(uuid.uuid4())
        run(conn, """
            INSERT INTO categories (id, chantier_id, nom, ordre, is_custom)
            VALUES (%s,%s,%s,%s,TRUE)
        """, (catid, d['chantierId'], d['nom'], ordre))
        conn.commit()
        return jsonify({'id': catid, 'chantierId': d['chantierId'], 'nom': d['nom'],
                        'ordre': ordre, 'isCustom': True,
                        'tasks': [], 'progress': 0, 'totalCount': 0,
                        'score': 0, 'maxScore': 0}), 201
    finally:
        conn.close()

@app.route('/api/categories/<catid>', methods=['DELETE'])
@login_required
def delete_categorie(catid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    try:
        run(conn, "DELETE FROM categories WHERE id=%s", (catid,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

# ── Tâches ────────────────────────────────────────────────────────────────────

def fmt_tache(t):
    return {'id': str(t['id']), 'categorieId': str(t['categorie_id']),
            'nom': t['nom'], 'etabli': t['etabli'],
            'envoye': t['envoye'], 'valide': t['valide']}

@app.route('/api/taches', methods=['POST'])
@login_required
def create_tache():
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    d = request.json
    tid = str(uuid.uuid4())
    conn = get_db()
    try:
        run(conn, """
            INSERT INTO taches (id, categorie_id, nom, etabli, envoye, valide)
            VALUES (%s,%s,%s,0,0,0)
        """, (tid, d['categorieId'], d['nom']))
        conn.commit()
        return jsonify({'id': tid, 'categorieId': d['categorieId'],
                        'nom': d['nom'], 'etabli': 0, 'envoye': 0, 'valide': 0}), 201
    finally:
        conn.close()

@app.route('/api/taches/<tid>', methods=['PUT'])
@login_required
def update_tache(tid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    d = request.json
    conn = get_db()
    try:
        fields = []
        vals   = []
        for k in ['nom','etabli','envoye','valide']:
            if k in d: fields.append(f'{k}=%s'); vals.append(d[k])
        if fields:
            run(conn, f"UPDATE taches SET {', '.join(fields)} WHERE id=%s", (*vals, tid))
            conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/taches/<tid>', methods=['DELETE'])
@login_required
def delete_tache(tid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    try:
        run(conn, "DELETE FROM taches WHERE id=%s", (tid,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

# ── Planning ──────────────────────────────────────────────────────────────────

def fmt_plan(r):
    return {'id': str(r['id']), 'chantierId': str(r['chantier_id']),
            'phase': r['phase'],
            'dateDebut': str(r['date_debut']), 'dateFin': str(r['date_fin'])}

@app.route('/api/planning/items')
@login_required
def get_planning_items():
    conn = get_db()
    try:
        return jsonify([fmt_plan(r) for r in qall(conn, "SELECT * FROM planning")])
    finally:
        conn.close()

@app.route('/api/planning/items', methods=['POST'])
@login_required
def create_planning_item():
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    d = request.json
    pid = str(uuid.uuid4())
    conn = get_db()
    try:
        run(conn, """
            INSERT INTO planning (id, chantier_id, phase, date_debut, date_fin)
            VALUES (%s,%s,%s,%s,%s)
        """, (pid, d['chantierId'], d['phase'], d['dateDebut'], d['dateFin']))
        conn.commit()
        return jsonify({'id': pid, 'chantierId': d['chantierId'], 'phase': d['phase'],
                        'dateDebut': d['dateDebut'], 'dateFin': d['dateFin']}), 201
    finally:
        conn.close()

@app.route('/api/planning/items/<pid>', methods=['PUT'])
@login_required
def update_planning_item(pid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    d = request.json
    conn = get_db()
    try:
        mapping = {'phase':'phase','dateDebut':'date_debut','dateFin':'date_fin'}
        fields, vals = [], []
        for k, col in mapping.items():
            if k in d: fields.append(f'{col}=%s'); vals.append(d[k])
        if fields:
            run(conn, f"UPDATE planning SET {', '.join(fields)} WHERE id=%s", (*vals, pid))
            conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/planning/items/<pid>', methods=['DELETE'])
@login_required
def delete_planning_item(pid):
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    conn = get_db()
    try:
        run(conn, "DELETE FROM planning WHERE id=%s", (pid,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

# ── Export / Import Excel ─────────────────────────────────────────────────────

@app.route('/api/export')
@login_required
def export_excel():
    wb = Workbook()
    conn = get_db()
    try:
        ws = wb.active; ws.title = 'Chantiers'
        ws.append(['id','nom','adresse','client','dateDebut','dateFin','commentaires'])
        for r in qall(conn, "SELECT * FROM chantiers ORDER BY nom"):
            ws.append([str(r['id']),r['nom'],r['adresse'],r['client'],
                       str(r['date_debut'] or ''),str(r['date_fin'] or ''),r['commentaires']])

        ws2 = wb.create_sheet('Categories')
        ws2.append(['id','chantierId','nom','ordre','isCustom'])
        for r in qall(conn, "SELECT * FROM categories ORDER BY ordre"):
            ws2.append([str(r['id']),str(r['chantier_id']),r['nom'],r['ordre'],r['is_custom']])

        ws3 = wb.create_sheet('Taches')
        ws3.append(['id','categorieId','nom','etabli','envoye','valide'])
        for r in qall(conn, "SELECT * FROM taches"):
            ws3.append([str(r['id']),str(r['categorie_id']),r['nom'],
                        r['etabli'],r['envoye'],r['valide']])

        ws4 = wb.create_sheet('Planning')
        ws4.append(['id','chantierId','phase','dateDebut','dateFin'])
        for r in qall(conn, "SELECT * FROM planning"):
            ws4.append([str(r['id']),str(r['chantier_id']),r['phase'],
                        str(r['date_debut']),str(r['date_fin'])])
    finally:
        conn.close()

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='chantiers.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/import', methods=['POST'])
@login_required
def import_excel():
    if readonly_check(): return jsonify({'error': 'Accès en lecture seule'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    try:
        wb  = openpyxl.load_workbook(request.files['file'])
        conn = get_db()
        try:
            run(conn, "DELETE FROM planning"); run(conn, "DELETE FROM taches")
            run(conn, "DELETE FROM categories"); run(conn, "DELETE FROM chantiers")

            def rows(name):
                if name not in wb.sheetnames: return
                ws = wb[name]; data = list(ws.iter_rows(values_only=True))
                if not data: return
                h = [str(x) for x in data[0]]
                for row in data[1:]:
                    if all(v is None for v in row): continue
                    yield {h[i]: row[i] for i in range(len(h))}

            for r in rows('Chantiers'):
                run(conn, """
                    INSERT INTO chantiers (id,nom,adresse,client,date_debut,date_fin,commentaires)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (r.get('id'), r.get('nom'), r.get('adresse'), r.get('client'),
                      r.get('dateDebut') or None, r.get('dateFin') or None, r.get('commentaires')))
            for r in rows('Categories'):
                run(conn, """
                    INSERT INTO categories (id,chantier_id,nom,ordre,is_custom)
                    VALUES (%s,%s,%s,%s,%s)
                """, (r.get('id'),r.get('chantierId'),r.get('nom'),
                      r.get('ordre',0), bool(r.get('isCustom',False))))
            for r in rows('Taches'):
                run(conn, """
                    INSERT INTO taches (id,categorie_id,nom,etabli,envoye,valide)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (r.get('id'),r.get('categorieId'),r.get('nom'),
                      r.get('etabli',0),r.get('envoye',0),r.get('valide',0)))
            for r in rows('Planning'):
                run(conn, """
                    INSERT INTO planning (id,chantier_id,phase,date_debut,date_fin)
                    VALUES (%s,%s,%s,%s,%s)
                """, (r.get('id'),r.get('chantierId'),r.get('phase'),
                      r.get('dateDebut'),r.get('dateFin')))
            conn.commit()
        finally:
            conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Lancement ──────────────────────────────────────────────────────────────────

init_db()

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
